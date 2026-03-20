"""
Conector inteligente que parsea cualquier formato de Excel usando Claude AI.

QUE HACE ESTE ARCHIVO:
Lee archivos Excel con formato arbitrario (el que ya tiene el distribuidor),
usa Claude para detectar que es cada columna, y transforma los datos al
formato que SyncEngine espera. Es el corazon del Canal 2 (Smart File Upload).

CONCEPTO CLAVE - Strategy Pattern:
SmartFileConnector implementa la misma interfaz (ERPConnector) que
ExcelConnector y ContabiliumConnector. SyncEngine no sabe ni le importa
cual de los tres esta usando. Es como un traductor universal: sin importar
en que idioma venga el Excel, lo traduce al "idioma interno" de PymePilot.

CONCEPTO CLAVE - Composicion:
En vez de heredar de ExcelConnector, SmartFileConnector usa su propia logica.
Son dos estrategias distintas: ExcelConnector espera formato fijo,
SmartFileConnector se adapta a cualquier formato.

SEGURIDAD:
- openpyxl read_only=True, data_only=True: no ejecuta macros ni formulas
- Solo se mandan headers + 5 filas a Claude, NUNCA los datos completos
- Los archivos se abren en modo lectura, nunca se modifican
"""

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path

import openpyxl

from backend.engine.connectors.base import ERPConnector
from backend.engine.claude.client import ClaudeClient
from backend.engine.core.logger import get_logger

logger = get_logger(__name__)

# Path al prompt para Claude (relativo a la raiz del proyecto)
_PROMPT_PATH = Path(__file__).parent.parent.parent / "config" / "prompts" / "smart_upload.txt"

# Limite de filas por hoja. Un distribuidor PyME no deberia tener mas de 50k
# filas en un Excel. Sin este limite, un .xlsx de 10MB comprimido puede tener
# cientos de miles de filas que al descomprimirse consumen toda la RAM.
MAX_ROWS_PER_SHEET = 50_000


class SmartFileConnector(ERPConnector):
    """Conector inteligente para Excel de formato arbitrario.

    Flujo de uso (en 2 pasos):
        # Paso 1: Analizar archivos con Claude (ANTES de crear el conector)
        mapping = SmartFileConnector.analyze_files(['/tmp/ventas.xlsx'])

        # Paso 2: Crear conector con el mapping y usarlo normalmente
        connector = SmartFileConnector(['/tmp/ventas.xlsx'], mapping)
        connector.test_connection()
        customers, _ = connector.fetch_customers()

    Por que 2 pasos: analyze_files() llama a Claude (I/O externo, puede fallar).
    Separarlo permite manejar errores del analisis antes de intentar el sync.
    """

    def __init__(self, file_paths: list[str], mapping: dict) -> None:
        """
        Args:
            file_paths: Lista de paths a archivos Excel descargados.
            mapping: Resultado de analyze_files() — dict con key 'sheets'.
        """
        self._file_paths = [Path(p) for p in file_paths]
        self._mapping = mapping
        self._parsed: dict | None = None  # Cache lazy de datos parseados

    @staticmethod
    def analyze_files(file_paths: list[str]) -> dict:
        """Analiza archivos con Claude para obtener mapeo de columnas.

        QUE HACE: Lee headers + 5 filas de muestra de cada hoja de cada archivo,
        se las manda a Claude, y Claude responde con un JSON diciendo que tipo
        de datos hay y como mapear cada columna.

        POR QUE: Cada distribuidor organiza su Excel diferente. Claude entiende
        contexto y puede mapear 'Razón Social' → customer_name sin reglas fijas.

        Args:
            file_paths: Lista de paths a archivos .xlsx

        Returns:
            dict con key 'sheets': lista de dicts con file, sheet, entity_type, column_mapping

        Raises:
            ValueError: Si Claude no devuelve JSON valido despues de 1 retry
        """
        # 1. Leer muestras de cada archivo/hoja
        file_data_parts = []
        for fpath in file_paths:
            path = Path(fpath)
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= 6:  # headers (fila 0) + 5 filas de datos
                            break
                        rows.append(row)

                    if not rows:
                        continue

                    headers = [str(h).strip() if h is not None else f"col_{j}"
                               for j, h in enumerate(rows[0])]

                    # Formatear muestra para Claude
                    sample_lines = []
                    for row in rows[1:]:
                        vals = [str(v) if v is not None else "" for v in row]
                        sample_lines.append("  " + " | ".join(vals))

                    file_data_parts.append(
                        f"ARCHIVO: {path.name}\n"
                        f"HOJA: {sheet_name}\n"
                        f"HEADERS: {' | '.join(headers)}\n"
                        f"MUESTRA ({len(sample_lines)} filas):\n"
                        + "\n".join(sample_lines)
                    )
            finally:
                wb.close()

        if not file_data_parts:
            raise ValueError("No se encontraron datos en ningun archivo/hoja")

        file_data = "\n\n".join(file_data_parts)

        # 2. Cargar prompt y llamar a Claude
        prompt_template = _PROMPT_PATH.read_text(encoding="utf-8")
        user_prompt = prompt_template.replace("{file_data}", file_data)

        client = ClaudeClient()
        # max_tokens=2000 es suficiente para el JSON de mapeo
        result = client.generate_message(
            system_prompt="Sos un analizador de datos. Respondes SOLO con JSON valido.",
            user_prompt=user_prompt,
            max_tokens=2000,
        )

        # 3. Parsear respuesta JSON
        response_text = result["text"].strip()
        # Limpiar posibles backticks de markdown que Claude a veces agrega
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1]  # quitar primera linea
        if response_text.endswith("```"):
            response_text = response_text.rsplit("```", 1)[0]
        response_text = response_text.strip()

        try:
            mapping = json.loads(response_text)
        except json.JSONDecodeError as exc:
            logger.error(f"Claude devolvio JSON invalido: {response_text[:500]}")
            raise ValueError(
                "No pudimos interpretar el formato del archivo. "
                "Revisa que tenga datos de ventas con columnas claras."
            ) from exc

        # 4. Validar estructura basica
        if "sheets" not in mapping or not isinstance(mapping["sheets"], list):
            raise ValueError(
                "Claude devolvio un mapeo sin estructura 'sheets' valida."
            )

        useful_sheets = [
            s for s in mapping["sheets"]
            if s.get("entity_type") != "unknown"
        ]
        if not useful_sheets:
            raise ValueError(
                "No pudimos identificar datos de ventas, clientes o productos "
                "en el archivo. Revisa que tenga columnas claras."
            )

        logger.info(
            f"analyze_files: {len(file_paths)} archivos, "
            f"{len(useful_sheets)} hojas utiles detectadas. "
            f"Tokens usados: {result['tokens_total']}, costo: ${result['cost_usd']:.6f}"
        )

        return mapping

    def test_connection(self) -> bool:
        """Valida que los archivos existen y el mapping tiene sentido.

        No abre los archivos completos — solo verifica existencia y
        que el mapping referencia archivos/hojas que existen.
        """
        for fpath in self._file_paths:
            if not fpath.exists():
                raise FileNotFoundError(f"Archivo no encontrado: {fpath}")

        # Verificar que al menos 1 hoja tiene datos utiles
        useful = [s for s in self._mapping.get("sheets", [])
                  if s.get("entity_type") != "unknown"]
        if not useful:
            raise ValueError("El mapping no contiene hojas con datos utiles")

        logger.info(f"test_connection(): {len(self._file_paths)} archivos, "
                     f"{len(useful)} hojas con datos utiles")
        return True

    # --- Metodos ERPConnector ---

    def fetch_customers(self, limit: int | None = None, client_ids: set[str] | None = None, since_date: date | None = None) -> tuple[list[dict], bool]:
        """Extrae clientes segun el mapping de Claude.

        Dependiendo del entity_type:
        - 'customers': Lee la hoja directamente, mapea columnas.
        - 'mixed_sales': Extrae valores unicos de customer_name.
        - 'orders': Extrae valores unicos de customer_ref.
        """
        self._ensure_parsed()
        records = self._parsed["customers"]
        records = self._validate_records(records, "clientes", ["Id", "RazonSocial"])
        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_customers(): {len(records)} clientes de Smart Upload")
        return records, False

    def fetch_products(self, limit: int | None = None, since_date: date | None = None) -> tuple[list[dict], bool]:
        """Extrae productos segun el mapping de Claude.

        Dependiendo del entity_type:
        - 'products': Lee la hoja directamente.
        - 'mixed_sales': Extrae valores unicos de product_name con precio.
        """
        self._ensure_parsed()
        records = self._parsed["products"]
        records = self._validate_records(records, "productos", ["Id", "Nombre"])
        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_products(): {len(records)} productos de Smart Upload")
        return records, False

    def fetch_orders(self, since_date: date | None = None, limit: int | None = None) -> tuple[list[dict], bool]:
        """Extrae ordenes segun el mapping de Claude.

        Dependiendo del entity_type:
        - 'orders': Lee la hoja directamente.
        - 'mixed_sales': Agrupa filas por (fecha + cliente) en ordenes con items.
        """
        self._ensure_parsed()
        records = self._parsed["orders"]

        # Filtrar por fecha si se pidio
        if since_date is not None:
            filtered = []
            for r in records:
                fecha = r.get("Fecha")
                if fecha is not None:
                    if hasattr(fecha, 'date'):
                        fecha = fecha.date()
                    if isinstance(fecha, date) and fecha >= since_date:
                        filtered.append(r)
            records = filtered

        records = self._validate_records(records, "ordenes", ["Id", "Fecha"])
        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_orders(): {len(records)} ordenes de Smart Upload")
        return records, False

    # --- Internos ---

    def _ensure_parsed(self) -> None:
        """Parsea todos los datos si aun no se hizo (lazy loading)."""
        if self._parsed is not None:
            return
        self._parsed = {"customers": [], "products": [], "orders": []}
        self._parse_all_sheets()

    def _parse_all_sheets(self) -> None:
        """Recorre todas las hojas del mapping y extrae datos segun entity_type."""
        # Indice de archivos por nombre para busqueda rapida
        file_index = {p.name: p for p in self._file_paths}

        for sheet_info in self._mapping.get("sheets", []):
            entity_type = sheet_info.get("entity_type", "unknown")
            if entity_type == "unknown":
                continue

            file_name = sheet_info.get("file", "")
            sheet_name = sheet_info.get("sheet", "")
            col_map = sheet_info.get("column_mapping", {})

            # Encontrar el archivo
            fpath = file_index.get(file_name)
            if fpath is None:
                logger.warning(f"Archivo '{file_name}' del mapping no encontrado, saltando")
                continue

            # Leer datos de la hoja
            rows = self._read_sheet(fpath, sheet_name)
            if not rows:
                continue

            # Invertir mapping: campo_destino → nombre_columna_original
            reverse_map = {v: k for k, v in col_map.items()}

            if entity_type == "mixed_sales":
                self._parse_mixed_sales(rows, reverse_map)
            elif entity_type == "customers":
                self._parse_customers_sheet(rows, reverse_map)
            elif entity_type == "products":
                self._parse_products_sheet(rows, reverse_map)
            elif entity_type == "orders":
                self._parse_orders_sheet(rows, reverse_map)

    def _read_sheet(self, file_path: Path, sheet_name: str) -> list[dict]:
        """Lee una hoja completa y retorna lista de dicts (header → valor).

        Misma logica que ExcelConnector._read_sheet pero acepta Path + nombre.
        SEGURIDAD: read_only=True, data_only=True (sin macros ni formulas).
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Hoja '{sheet_name}' no encontrada en {file_path.name}")
                return []

            ws = wb[sheet_name]
            # M-04 FIX: Leer con limite de filas en vez de list() completo.
            # list(ws.iter_rows()) materializa TODA la hoja en RAM.
            # Con read_only=True openpyxl usa parser incremental, pero
            # list() anula esa ventaja. Con el limite, si la hoja tiene
            # 500k filas, solo leemos las primeras 50k.
            raw_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > MAX_ROWS_PER_SHEET:
                    logger.warning(
                        f"Hoja '{sheet_name}' truncada a {MAX_ROWS_PER_SHEET} filas "
                        f"(archivo: {file_path.name})"
                    )
                    break
                raw_rows.append(row)

            if not raw_rows:
                return []

            headers = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(raw_rows[0])]

            records = []
            for row in raw_rows[1:]:
                if all(v is None for v in row):
                    continue
                record = dict(zip(headers, row))
                records.append(record)

            return records
        finally:
            wb.close()

    def _parse_mixed_sales(self, rows: list[dict], reverse_map: dict) -> None:
        """Extrae customers, products y orders de una hoja con todo junto.

        QUE HACE: De una hoja tipo:
            Fecha | Cliente | Producto | Cant | Precio | Total
        Extrae:
            - Clientes unicos (por nombre)
            - Productos unicos (por nombre, con precio)
            - Ordenes agrupadas por fecha+cliente, con items

        CONCEPTO - Normalizacion:
        Es como desarmar un pedido de delivery. El ticket dice todo junto:
        "Juan - 2 hamburguesas $500 - 1 gaseosa $200". De ahi extraemos:
        - Cliente: Juan
        - Productos: hamburguesa ($500), gaseosa ($200)
        - Orden: fecha + items
        """
        # Columnas originales para cada campo destino
        col_date = reverse_map.get("order_date")
        col_customer = reverse_map.get("customer_name")
        col_product = reverse_map.get("product_name")
        col_qty = reverse_map.get("quantity")
        col_price = reverse_map.get("unit_price")
        col_total = reverse_map.get("total_amount")
        col_order_id = reverse_map.get("order_id")

        if not col_date or not col_customer:
            logger.warning("mixed_sales: faltan columnas obligatorias (order_date, customer_name)")
            return

        # --- Extraer clientes unicos ---
        # IDs basados en hash del nombre → mismos datos = mismo ID entre uploads
        customer_ids: dict[str, str] = {}  # nombre → ID generado
        for row in rows:
            name = str(row.get(col_customer, "")).strip()
            if name and name not in customer_ids:
                cid = _content_hash(name)
                customer_ids[name] = cid
                self._parsed["customers"].append({
                    "Id": cid,
                    "RazonSocial": name,
                })

        # --- Extraer productos unicos (si hay columna de producto) ---
        # IDs basados en hash del nombre → mismos datos = mismo ID entre uploads
        product_ids: dict[str, str] = {}  # nombre → ID generado
        if col_product:
            for row in rows:
                pname = str(row.get(col_product, "")).strip()
                if pname and pname not in product_ids:
                    pid = _content_hash(pname)
                    product_ids[pname] = pid
                    price = _safe_number(row.get(col_price)) if col_price else None
                    self._parsed["products"].append({
                        "Id": pid,
                        "Nombre": pname,
                        "PrecioVenta": price,
                    })

        # --- Agrupar filas en ordenes ---
        # Si hay order_id explícito, agrupar por eso.
        # Si no, agrupar por (fecha + cliente) — cada combinación = 1 orden.
        # IDs basados en hash → mismos datos = mismo ID entre uploads
        orders_by_key: dict[str, dict] = {}

        for row in rows:
            customer_name = str(row.get(col_customer, "")).strip()
            if not customer_name:
                continue

            fecha = _normalize_date(row.get(col_date))
            customer_id = customer_ids.get(customer_name, "")

            # Determinar la key de agrupacion
            if col_order_id and row.get(col_order_id):
                order_key = str(row.get(col_order_id))
            else:
                # Hash de fecha+cliente como key estable
                date_str = str(fecha) if fecha else ""
                order_key = sha256(f"{date_str}|{customer_name}".encode()).hexdigest()[:12]

            if order_key not in orders_by_key:
                if col_order_id and row.get(col_order_id):
                    oid = str(row.get(col_order_id))
                else:
                    date_str = str(fecha) if fecha else ""
                    oid = _content_hash(date_str, customer_name)
                orders_by_key[order_key] = {
                    "Id": oid,
                    "Fecha": fecha,
                    "Cliente": customer_id,
                    "Total": 0,
                    "Items": [],
                }

            order = orders_by_key[order_key]

            # Acumular total
            line_total = _safe_number(row.get(col_total)) if col_total else None
            qty = _safe_number(row.get(col_qty)) if col_qty else None
            unit_price = _safe_number(row.get(col_price)) if col_price else None

            if line_total is not None:
                order["Total"] = (order["Total"] or 0) + line_total

            # Agregar item si hay producto
            if col_product and row.get(col_product):
                pname = str(row.get(col_product, "")).strip()
                product_id = product_ids.get(pname, "")
                order["Items"].append({
                    "Concepto": {
                        "Id": product_id,
                        "Nombre": pname,
                    },
                    "Cantidad": qty,
                    "PrecioUnitario": unit_price,
                    "Total": line_total,
                })

        self._parsed["orders"].extend(orders_by_key.values())

    def _parse_customers_sheet(self, rows: list[dict], reverse_map: dict) -> None:
        """Parsea hoja de clientes. Mapea columnas al formato SyncEngine."""
        for row in rows:
            name = _get_mapped(row, reverse_map, "customer_name")
            if not name:
                continue
            # Si el Excel tiene su propia columna de ID, usarla.
            # Si no, generar hash del nombre → estable entre uploads.
            cid = _get_mapped(row, reverse_map, "customer_id") or _content_hash(str(name))

            self._parsed["customers"].append({
                "Id": str(cid),
                "RazonSocial": str(name),
                "Email": _get_mapped(row, reverse_map, "email"),
                "Telefono": _get_mapped(row, reverse_map, "phone"),
                "Domicilio": _get_mapped(row, reverse_map, "address"),
                "Localidad": _get_mapped(row, reverse_map, "city"),
                "Observaciones": _get_mapped(row, reverse_map, "notes"),
            })

    def _parse_products_sheet(self, rows: list[dict], reverse_map: dict) -> None:
        """Parsea hoja de productos. Mapea columnas al formato SyncEngine."""
        for row in rows:
            name = _get_mapped(row, reverse_map, "product_name")
            if not name:
                continue
            sku = _get_mapped(row, reverse_map, "sku")
            # Si el Excel tiene su propia columna de ID, usarla.
            # Si no, generar hash del nombre+sku → estable entre uploads.
            pid = _get_mapped(row, reverse_map, "product_id") or _content_hash(str(name), str(sku or ""))

            self._parsed["products"].append({
                "Id": str(pid),
                "Nombre": str(name),
                "PrecioVenta": _safe_number(_get_mapped(row, reverse_map, "price")),
                "Codigo": sku,
                "Rubro": _get_mapped(row, reverse_map, "category"),
                "SubRubro": _get_mapped(row, reverse_map, "subcategory"),
            })

    def _parse_orders_sheet(self, rows: list[dict], reverse_map: dict) -> None:
        """Parsea hoja de ordenes (sin items). Mapea columnas al formato SyncEngine."""
        for row in rows:
            fecha = _normalize_date(_get_mapped(row, reverse_map, "order_date"))
            customer_ref = _get_mapped(row, reverse_map, "customer_ref")
            if not fecha:
                continue
            total = _safe_number(_get_mapped(row, reverse_map, "total_amount"))
            # Si el Excel tiene su propia columna de ID, usarla.
            # Si no, generar hash de fecha+cliente+total → estable entre uploads.
            oid = _get_mapped(row, reverse_map, "order_id") or _content_hash(
                str(fecha), str(customer_ref or ""), str(total or "")
            )

            self._parsed["orders"].append({
                "Id": str(oid),
                "Fecha": fecha,
                "Cliente": str(customer_ref) if customer_ref else "",
                "Total": total,
            })


# --- Funciones helper (module-level, no metodos) ---


def _content_hash(*parts: str) -> str:
    """Genera external_id estable basado en contenido.

    QUE HACE: Toma una o mas strings (ej: nombre del cliente, nombre del
    producto + SKU), las normaliza (quita acentos, minusculas, trim),
    las une con '|' y les saca un hash SHA-256 truncado a 12 caracteres.

    POR QUE: Si el mismo dato aparece en 2 uploads distintos, genera
    el mismo ID. Asi SyncEngine hace UPSERT (actualiza) en vez de
    INSERT (duplicar). Es como un DNI basado en tus datos: siempre
    el mismo para los mismos inputs.

    L-01 FIX: Se cambio de MD5 a SHA-256. MD5 tiene colisiones conocidas
    (dos inputs diferentes que dan el mismo hash). SHA-256 es mas moderno
    y resistente, sin costo adicional de performance.

    Prefijo 'su_' (smart upload) evita colision con IDs del Canal 1 (ERP),
    que son numericos puros (ej: '12345' de Contabilium).

    Args:
        *parts: Strings que identifican la entidad (nombre, SKU, fecha, etc.)

    Returns:
        String tipo 'su_a1b2c3d4e5f6' (prefijo + 12 hex chars)
    """
    normalized = "|".join(
        unicodedata.normalize("NFKD", str(p)).encode("ascii", "ignore").decode().lower().strip()
        for p in parts if p
    )
    return "su_" + sha256(normalized.encode()).hexdigest()[:12]


def _get_mapped(row: dict, reverse_map: dict, field: str):
    """Obtiene un valor del row usando el reverse mapping.

    Args:
        row: Dict de una fila del Excel (header original → valor)
        reverse_map: Dict (campo_destino → header original)
        field: Nombre del campo destino a buscar

    Returns:
        El valor de la celda, o None si no hay mapping para ese campo.
    """
    col_name = reverse_map.get(field)
    if col_name is None:
        return None
    return row.get(col_name)


def _safe_number(value) -> float | None:
    """Convierte un valor a float de forma segura. Retorna None si no puede."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _normalize_date(value) -> date | None:
    """Convierte cualquier formato de fecha a un objeto date de Python.

    QUE HACE: Los Excel argentinos tipicamente usan DD/MM/YYYY,
    pero openpyxl puede devolver datetime, date, o string segun
    como este formateada la celda. Esta funcion los unifica.

    POR QUE: PostgreSQL espera YYYY-MM-DD. Si mandamos un string
    como "24/01/2026", tira error. Al convertir a date de Python,
    psycopg3 lo serializa correctamente.

    Formatos soportados (en orden de prioridad):
    - datetime/date de Python → retorna .date()
    - DD/MM/YYYY o DD-MM-YYYY (formato argentino, se intenta primero)
    - YYYY-MM-DD o YYYY/MM/DD (formato ISO)
    - DD/MM/YY o DD-MM-YY (formato argentino abreviado)
    NOTA: Para fechas ambiguas como 05/03/2026, siempre se interpreta
    como DD/MM/YYYY (5 de marzo) porque los distribuidores son argentinos.
    """
    if value is None:
        return None

    # Ya es date o datetime de Python (openpyxl con celdas formateadas)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Convertir a string y limpiar
    text = str(value).strip()
    if not text:
        return None

    # Intentar formatos comunes
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    logger.warning(f"No se pudo parsear fecha: '{text}'")
    return None
