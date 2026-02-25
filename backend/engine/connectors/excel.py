"""
Conector para archivos Excel (.xlsx) como fuente de datos alternativa.

QUE HACE ESTE ARCHIVO:
Lee datos de clientes, productos y ventas desde un archivo Excel.
Es el fallback para distribuidores que no tienen API en su ERP.

CONCEPTO CLAVE - Strategy Pattern:
Tanto ContabiliumConnector como ExcelConnector implementan la misma
interfaz (ERPConnector). El codigo que usa el conector no sabe ni le
importa de donde vienen los datos.

CONCEPTO CLAVE - Adapter Pattern (mapeo de campos):
El SyncEngine espera datos con nombres de campos al estilo Contabilium
(PrecioVenta, Domicilio, Rubro, etc). Pero el Excel usa nombres amigables
(Precio, Direccion, Categoria). Este conector traduce entre ambos formatos,
como un enchufe adaptador de viaje.

SEGURIDAD:
- read_only=True: abre en modo solo lectura (no modifica el archivo fuente)
- data_only=True: lee valores calculados, NO evalua formulas.
  Esto deshabilita evaluacion de formulas y macros, cerrando un
  vector de ataque completo (archivos Excel maliciosos).
"""

from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

from backend.engine.connectors.base import ERPConnector
from backend.engine.core.logger import get_logger

logger = get_logger(__name__)

# Limite de filas por hoja. Mismo valor que SmartFileConnector.
# Protege contra archivos enormes que consumirian toda la RAM.
MAX_ROWS_PER_SHEET = 50_000

# Columnas minimas requeridas por hoja.
# "Id" es obligatorio en todas: sin el, external_id queda vacio y el
# UNIQUE constraint (tenant_id, external_id) falla en el segundo registro.
_REQUIRED_COLUMNS = {
    "Clientes": {"Id", "Nombre"},  # + al menos Email O Telefono
    "Productos": {"Id", "Nombre", "Precio"},
    "Ventas": {"Id", "Fecha", "Cliente", "Total"},
}

# Mapeo: nombre amigable en Excel → nombre que SyncEngine espera.
# Solo los campos que difieren. Los que ya coinciden (Id, Nombre,
# Email, Telefono, Fecha, Total, Cliente) no necesitan traduccion.
_FIELD_MAP_CUSTOMERS = {
    "Nombre": "RazonSocial",
    "Direccion": "Domicilio",
    "Ciudad": "Localidad",
    "Notas": "Observaciones",
}

_FIELD_MAP_PRODUCTS = {
    "Precio": "PrecioVenta",
    "SKU": "Codigo",
    "Categoria": "Rubro",
    "Subcategoria": "SubRubro",
}


class ExcelConnector(ERPConnector):
    """Conector de solo lectura para archivos Excel (.xlsx).

    Uso:
        connector = ExcelConnector('/ruta/al/archivo.xlsx')
        connector.test_connection()  # valida estructura
        customers, truncated = connector.fetch_customers()

    Hojas esperadas:
        - Clientes: Id, Nombre, Email, Telefono, [Direccion, Ciudad, Notas]
        - Productos: Id, Nombre, Precio, [SKU, Categoria, Subcategoria]
        - Ventas: Id, Fecha, Cliente (external_id del cliente), Total
        - Items (opcional): OrdenId, ProductoId, Cantidad, PrecioUnitario, Total
    """

    def __init__(self, file_path: str) -> None:
        self._file_path = Path(file_path)

    def test_connection(self) -> bool:
        """Validacion completa del archivo Excel.

        1. Verifica que el archivo existe
        2. Verifica que tiene las hojas esperadas: "Clientes", "Productos", "Ventas"
        3. Verifica columnas minimas requeridas por hoja
        """
        # 1. Verificar existencia
        if not self._file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self._file_path}")

        # 2. Verificar hojas
        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            expected_sheets = {"Clientes", "Productos", "Ventas"}
            actual_sheets = set(wb.sheetnames)
            missing_sheets = expected_sheets - actual_sheets

            if missing_sheets:
                raise ValueError(
                    f"Hojas faltantes: {sorted(missing_sheets)}. "
                    f"Hojas encontradas: {sorted(actual_sheets)}"
                )

            # 3. Verificar columnas minimas por hoja
            for sheet_name, required_cols in _REQUIRED_COLUMNS.items():
                ws = wb[sheet_name]
                # Primera fila = headers
                headers = set()
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    if cell.value is not None:
                        headers.add(str(cell.value).strip())

                missing_cols = required_cols - headers
                if missing_cols:
                    raise ValueError(
                        f"Hoja {sheet_name}: columnas faltantes: {sorted(missing_cols)}. "
                        f"Columnas encontradas: {sorted(headers)}"
                    )

            # Validacion especial para Clientes: necesita Email O Telefono
            ws_clientes = wb["Clientes"]
            headers_clientes = set()
            for cell in next(ws_clientes.iter_rows(min_row=1, max_row=1)):
                if cell.value is not None:
                    headers_clientes.add(str(cell.value).strip())
            if "Email" not in headers_clientes and "Telefono" not in headers_clientes:
                raise ValueError(
                    f"Hoja Clientes: necesita al menos 'Email' o 'Telefono'. "
                    f"Columnas encontradas: {sorted(headers_clientes)}"
                )

            # Info sobre hoja Items (opcional, no requerida)
            if "Items" in actual_sheets:
                logger.info("test_connection(): hoja 'Items' detectada (detalle de ventas)")
            else:
                logger.info("test_connection(): sin hoja 'Items' (ordenes sin detalle de productos)")

        finally:
            wb.close()

        logger.info(f"test_connection(): archivo Excel valido: {self._file_path.name}")
        return True

    def _read_sheet(self, sheet_name: str, required: bool = True) -> list[dict]:
        """Lee una hoja completa y retorna lista de dicts (header → valor).

        Args:
            sheet_name: Nombre de la hoja a leer.
            required: Si False y la hoja no existe, retorna lista vacia
                      en vez de lanzar excepcion. Util para hojas opcionales.
        """
        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            if not required and sheet_name not in wb.sheetnames:
                return []

            ws = wb[sheet_name]
            # Leer con limite de filas para proteger contra archivos enormes.
            # Mismo patron que SmartFileConnector._read_sheet().
            raw_rows: list[tuple] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > MAX_ROWS_PER_SHEET:
                    logger.warning(
                        f"Hoja '{sheet_name}' truncada a {MAX_ROWS_PER_SHEET} filas"
                    )
                    break
                raw_rows.append(row)

            if not raw_rows:
                return []

            headers = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(raw_rows[0])]

            records = []
            for row in raw_rows[1:]:
                # Saltar filas completamente vacias
                if all(v is None for v in row):
                    continue
                record = dict(zip(headers, row))
                records.append(record)

            return records
        finally:
            wb.close()

    @staticmethod
    def _apply_field_map(records: list[dict], field_map: dict[str, str]) -> list[dict]:
        """Renombra claves de cada dict segun el mapeo.

        No modifica los dicts originales (crea copias).
        Claves que no estan en el mapeo se mantienen tal cual.
        """
        mapped = []
        for record in records:
            new_record = {}
            for key, value in record.items():
                new_key = field_map.get(key, key)
                new_record[new_key] = value
            mapped.append(new_record)
        return mapped

    def fetch_customers(self, limit: int | None = None, client_ids: set[str] | None = None) -> tuple[list[dict], bool]:
        """Lee hoja 'Clientes' y mapea campos al formato SyncEngine."""
        records = self._read_sheet("Clientes")
        records = self._apply_field_map(records, _FIELD_MAP_CUSTOMERS)
        records = self._validate_records(records, "clientes", ["Id", "RazonSocial"])
        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_customers(): {len(records)} clientes validos de Excel")
        return records, False

    def fetch_products(self, limit: int | None = None) -> tuple[list[dict], bool]:
        """Lee hoja 'Productos' y mapea campos al formato SyncEngine."""
        records = self._read_sheet("Productos")
        records = self._apply_field_map(records, _FIELD_MAP_PRODUCTS)
        records = self._validate_records(records, "productos", ["Id", "Nombre"])
        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_products(): {len(records)} productos validos de Excel")
        return records, False

    def fetch_orders(self, since_date: date | None = None, limit: int | None = None) -> tuple[list[dict], bool]:
        """Lee hoja 'Ventas' + 'Items' (opcional) y arma estructura para SyncEngine.

        Si existe la hoja 'Items', agrupa los items por OrdenId y los adjunta
        a cada orden en el formato que SyncEngine espera (compatible Contabilium):
            order['Items'] = [{'Concepto': {'Id': ...}, 'Cantidad': ..., ...}]
        """
        records = self._read_sheet("Ventas")

        # Leer items si la hoja existe (required=False → lista vacia si no hay)
        items_raw = self._read_sheet("Items", required=False)
        if items_raw:
            # Agrupar items por OrdenId
            items_by_order: dict[str, list[dict]] = defaultdict(list)
            for item in items_raw:
                orden_id = str(item.get("OrdenId", ""))
                if orden_id:
                    items_by_order[orden_id].append({
                        "Concepto": {
                            "Id": str(item.get("ProductoId", "")),
                            "Nombre": str(item.get("NombreProducto", "")),
                        },
                        "Cantidad": item.get("Cantidad"),
                        "PrecioUnitario": item.get("PrecioUnitario"),
                        "Total": item.get("Total"),
                    })

            # Adjuntar items a cada orden
            for record in records:
                order_id = str(record.get("Id", ""))
                if order_id in items_by_order:
                    record["Items"] = items_by_order[order_id]

            logger.info(f"fetch_orders(): {len(items_raw)} items leidos de hoja 'Items'")

        # Validar campos obligatorios antes de filtrar
        records = self._validate_records(records, "ordenes", ["Id", "Fecha"])

        # Filtrar por fecha si se pidio
        if since_date is not None:
            filtered = []
            for r in records:
                fecha = r.get("Fecha")
                if fecha is not None:
                    # openpyxl puede devolver datetime o date directamente
                    if hasattr(fecha, 'date'):
                        fecha = fecha.date()
                    if isinstance(fecha, date) and fecha >= since_date:
                        filtered.append(r)
            records = filtered

        if limit is not None:
            records = records[:limit]
        logger.info(f"fetch_orders(): {len(records)} ordenes leidas de Excel")
        return records, False
