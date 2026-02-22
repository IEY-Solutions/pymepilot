#!/usr/bin/env python3
"""
Genera un archivo Excel de prueba para testear el sync via ExcelConnector.

QUE HACE ESTE SCRIPT:
Crea un archivo .xlsx con 4 hojas (Clientes, Productos, Ventas, Items)
con datos ficticios pero realistas de IEY (accesorios MagSafe).

POR QUE:
Contabilium esta bloqueado por Cloudflare (ticket abierto). Este Excel
nos permite validar todo el pipeline de sync sin depender de la API.

DATOS:
- 20 clientes (external_ids 3001-3020, para no pisar seed 1001-1020)
- 10 productos MagSafe (external_ids 4001-4010, para no pisar seed 2001-2010)
- 30 ventas con 65 items (external_ids 5001-5030)
- Cada venta referencia un cliente y tiene 1-4 items

USO:
  python backend/scripts/generate_test_excel.py
  # Genera: data/test/iey_test_data.xlsx
"""

import os
import sys
from datetime import date, timedelta
from random import Random

# Agregar raiz del proyecto al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import openpyxl


def main() -> None:
    # Seed fijo para reproducibilidad (mismos datos cada vez que se ejecuta)
    rng = Random(42)

    wb = openpyxl.Workbook()

    # =========================================================
    # HOJA 1: Clientes (20 clientes ficticios de IEY)
    # =========================================================
    ws_clientes = wb.active
    ws_clientes.title = "Clientes"
    ws_clientes.append(["Id", "Nombre", "Email", "Telefono", "Direccion", "Ciudad", "Notas"])

    clientes = [
        (3001, "Celulares Martinez", "martinez@cel.com.ar", "11-4001-0001", "Av. Corrientes 1234", "CABA", "Cliente desde 2024"),
        (3002, "TechMobile Cordoba", "info@techmobile.com.ar", "351-400-0002", "Av. Colon 890", "Cordoba", "Compra mensual"),
        (3003, "FundaShop Online", "ventas@fundashop.com.ar", "11-4001-0003", "Lavalle 567", "CABA", "E-commerce"),
        (3004, "Accesorios del Sur", "contacto@accsur.com.ar", "11-4001-0004", "Av. Mitre 345", "Avellaneda", ""),
        (3005, "MobileParts Rosario", "compras@mobileparts.com.ar", "341-400-0005", "San Martin 1100", "Rosario", "Distribuidor regional"),
        (3006, "CasePro La Plata", "info@casepro.com.ar", "221-400-0006", "Calle 7 N 890", "La Plata", ""),
        (3007, "iStore Mendoza", "ventas@istoremza.com.ar", "261-400-0007", "San Martin 456", "Mendoza", "Solo productos Apple"),
        (3008, "GadgetBox Tucuman", "gadgetbox@gmail.com", "381-400-0008", "24 de Septiembre 200", "Tucuman", ""),
        (3009, "PhoneZone CABA", "admin@phonezone.com.ar", "11-4001-0009", "Florida 300", "CABA", "Local en galeria"),
        (3010, "SmartCase MDP", "smartcase@hotmail.com", "223-400-0010", "San Martin 2500", "Mar del Plata", "Temporada alta verano"),
        (3011, "TechPoint Salta", "techpoint@gmail.com", "387-400-0011", "Caseros 600", "Salta", ""),
        (3012, "Digital Express", "info@digitalexpress.com.ar", "11-4001-0012", "Av. Santa Fe 1800", "CABA", "Envio gratis CABA"),
        (3013, "MagStore Neuquen", "magstore@outlook.com", "299-400-0013", "Av. Argentina 300", "Neuquen", ""),
        (3014, "ProCell Santa Fe", "procell@gmail.com", "342-400-0014", "San Martin 700", "Santa Fe", ""),
        (3015, "AppleZone Bahia", "applezone@gmail.com", "291-400-0015", "Alsina 150", "Bahia Blanca", "Unico Apple en la zona"),
        (3016, "Funda Express", "fundaexpress@yahoo.com.ar", "11-4001-0016", "Av. Rivadavia 5000", "CABA", "Mayorista"),
        (3017, "CelMax Parana", "celmax@gmail.com", "343-400-0017", "Urquiza 800", "Parana", ""),
        (3018, "TechShield SJ", "techshield@gmail.com", "264-400-0018", "Av. Libertador 400", "San Juan", ""),
        (3019, "MobileArg Resistencia", "mobilarg@gmail.com", "362-400-0019", "Junin 500", "Resistencia", ""),
        (3020, "FundaMaster", "fundamaster@gmail.com", "11-4001-0020", "Av. Cabildo 2000", "CABA", "Solo fundas premium"),
    ]

    for c in clientes:
        ws_clientes.append(list(c))

    # =========================================================
    # HOJA 2: Productos (10 productos MagSafe de IEY)
    # =========================================================
    ws_productos = wb.create_sheet("Productos")
    ws_productos.append(["Id", "SKU", "Nombre", "Precio", "Categoria", "Subcategoria"])

    productos = [
        (4001, "IEY-MS-CASE14", "Funda MagSafe iPhone 14", 15000, "Fundas", "iPhone 14"),
        (4002, "IEY-MS-CASE15", "Funda MagSafe iPhone 15", 18000, "Fundas", "iPhone 15"),
        (4003, "IEY-MS-CASE16", "Funda MagSafe iPhone 16", 22000, "Fundas", "iPhone 16"),
        (4004, "IEY-MS-CHG01", "Cargador MagSafe 15W", 25000, "Cargadores", "Inalambrico"),
        (4005, "IEY-MS-WAL01", "Billetera MagSafe Cuero", 20000, "Accesorios", "Billeteras"),
        (4006, "IEY-MS-STD01", "Soporte MagSafe Escritorio", 12000, "Soportes", "Escritorio"),
        (4007, "IEY-MS-CAR01", "Soporte MagSafe Auto", 16000, "Soportes", "Vehicular"),
        (4008, "IEY-MS-BAT01", "Battery Pack MagSafe", 30000, "Baterias", "Portatil"),
        (4009, "IEY-MS-GRP01", "PopGrip MagSafe", 8000, "Accesorios", "Grips"),
        (4010, "IEY-MS-VT15", "Vidrio Templado iPhone 15", 5000, "Protectores", "Vidrio"),
    ]

    for p in productos:
        ws_productos.append(list(p))

    # =========================================================
    # HOJA 3: Ventas (30 ordenes distribuidas en 3 meses)
    # =========================================================
    ws_ventas = wb.create_sheet("Ventas")
    ws_ventas.append(["Id", "Fecha", "Cliente", "Total"])

    # =========================================================
    # HOJA 4: Items (detalle de cada venta)
    # =========================================================
    ws_items = wb.create_sheet("Items")
    ws_items.append(["OrdenId", "ProductoId", "NombreProducto", "Cantidad", "PrecioUnitario", "Total"])

    # Generar 30 ventas con items aleatorios pero realistas
    client_ids = [c[0] for c in clientes]
    product_list = [(p[0], p[2], p[3]) for p in productos]  # (id, nombre, precio)

    base_date = date(2025, 12, 1)  # 3 meses de historial

    all_items = []
    for order_idx in range(30):
        order_id = 5001 + order_idx
        # Fecha aleatoria en los ultimos 3 meses
        order_date = base_date + timedelta(days=rng.randint(0, 82))
        # Cliente aleatorio
        client_id = rng.choice(client_ids)

        # 1 a 4 items por orden
        num_items = rng.randint(1, 4)
        # Elegir productos sin repetir dentro de la misma orden
        chosen_products = rng.sample(product_list, min(num_items, len(product_list)))

        order_total = 0
        for prod_id, prod_name, prod_price in chosen_products:
            qty = rng.randint(5, 50)
            line_total = qty * prod_price
            order_total += line_total
            all_items.append((order_id, prod_id, prod_name, qty, prod_price, line_total))

        ws_ventas.append([order_id, order_date, client_id, order_total])

    for item in all_items:
        ws_items.append(list(item))

    # =========================================================
    # GUARDAR
    # =========================================================
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "data", "test", "iey_test_data.xlsx"
    )
    wb.save(output_path)
    print(f"Excel generado: {output_path}")
    print(f"  Clientes: {len(clientes)}")
    print(f"  Productos: {len(productos)}")
    print(f"  Ventas: 30")
    print(f"  Items: {len(all_items)}")


if __name__ == '__main__':
    main()
